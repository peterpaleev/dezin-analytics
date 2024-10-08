import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Initialize the workbook
wb = Workbook()

# Remove the default sheet
default_sheet = wb.active
wb.remove(default_sheet)

# Data: Raw Materials and Prices
raw_materials_data = [
    {"Name": "Вода обессоленная", "Price": ""},
    {"Name": "Изопропанол", "Price": 168.5, "Unit": "кг"},
    {"Name": "Дезин", "Price": ""},
    {"Name": "Пэг 400", "Price": ""},
    {"Name": "Раствор красителя синий патентованный VE 131", "Price": ""},
    {"Name": "Эмал 270 d", "Price": ""},
    {"Name": "Сульфоэтоксилат", "Price": ""},
    {"Name": "Hansanol", "Price": ""},
    {"Name": "лаурил сульфат натрия", "Price": 235, "Unit": "кг"},
    {"Name": "эмерсенс АМ 8025", "Price": ""},
    {"Name": "кокосовый диэтаноламид", "Price": ""},
    {"Name": "глицерин", "Price": ""},
    {"Name": "пропиленгликоль", "Price": ""},
    {"Name": "алантоин 3b", "Price": ""},
    {"Name": "тетранил У", "Price": ""},
    {"Name": "кислота лимонная", "Price": ""},
    {"Name": "микрокер IT", "Price": ""},
    {"Name": "микрокер DH", "Price": ""},
    {"Name": "отдушка calendula 522", "Price": ""},
    {"Name": "дельтаглюконолактон", "Price": 225, "Unit": "кг"},
    {"Name": "хлоргексидин основание", "Price": 32.25, "Unit": "евро/кг"},
    {"Name": "пропанол 1/н-пропанол", "Price": ""},
    {"Name": "бардап 26", "Price": ""},
    {"Name": "вантоцил TG", "Price": ""},
    {"Name": "ультра отдушка 10010404", "Price": 307, "Unit": "кг"},
    {"Name": "арквад MCB -50", "Price": ""},
    {"Name": "1,2-пропиленгликоль", "Price": ""},
    {"Name": "краситель сансет желтый", "Price": ""},
    {"Name": "бардак 22 (Арквад 2.10-50", "Price": ""},
    {"Name": "барлокс 12", "Price": ""},
    {"Name": "лонзобак", "Price": ""},
    {"Name": "триамин", "Price": ""},
    {"Name": "Р-р красителя Блю Азур", "Price": ""},
    {"Name": "Дуванол РМ", "Price": ""},
    {"Name": "Барлокс 12", "Price": ""},
    {"Name": "краситель сансет е", "Price": ""},
    {"Name": "тартразин Е 102", "Price": ""},
    {"Name": "черный блестящий Е 151", "Price": ""},
    {"Name": "дезин", "Price": ""}
]

# Create Raw Materials sheet
rm_sheet = wb.create_sheet(title="Raw Materials")

# Headers
rm_sheet['A1'] = "Raw Material"
rm_sheet['B1'] = "Price"
rm_sheet['C1'] = "Unit"

# Populate Raw Materials data
for idx, material in enumerate(raw_materials_data, start=2):
    rm_sheet[f"A{idx}"] = material["Name"]
    rm_sheet[f"B{idx}"] = material["Price"] if material["Price"] != "" else ""
    rm_sheet[f"C{idx}"] = material.get("Unit", "")

# Define Products and their BOM (Bill of Materials)
products = {
    "Product A": {
        "Components": {
            "Изопропанол": 0.5,
            "лаурил сульфат натрия": 0.2,
            "дельтаглюконолактон": 0.1
        }
    },
    "Product B": {
        "Components": {
            "хлоргексидин основание": 0.05,
            "ультра отдушка 10010404": 0.02,
            "глицерин": 0.1
        }
    },
    # Add more products and their components as needed
}

# Create a sheet for the Production Cost Calculator
pcc_sheet = wb.create_sheet(title="Production Cost Calculator")

# Headers for Production Cost Calculator
pcc_headers = ["Product", "Quantity to Produce", "Total Cost", "Cost per Unit"]
pcc_sheet.append(pcc_headers)

# Start from the second row
row_idx = 2

for product_name, product_info in products.items():
    pcc_sheet.cell(row=row_idx, column=1).value = product_name
    # Quantity to Produce (you can set default values or leave them blank for input)
    pcc_sheet.cell(row=row_idx, column=2).value = 1000  # Default quantity
    # Formulas will be added later
    row_idx += 1

# Create a sheet for each product's BOM and calculations
for product_name, product_info in products.items():
    bom_sheet = wb.create_sheet(title=product_name)
    bom_sheet['A1'] = "Raw Material"
    bom_sheet['B1'] = "Quantity per Unit Product"
    bom_sheet['C1'] = "Unit"
    bom_sheet['D1'] = "Total Quantity Required"
    bom_sheet['E1'] = "Unit Price"
    bom_sheet['F1'] = "Total Cost"

    components = product_info["Components"]
    bom_row_idx = 2

    for material_name, qty_per_unit in components.items():
        # Raw Material Name
        bom_sheet.cell(row=bom_row_idx, column=1).value = material_name
        # Quantity per Unit Product
        bom_sheet.cell(row=bom_row_idx, column=2).value = qty_per_unit
        # Unit - Fetch from Raw Materials sheet
        for rm_row in range(2, len(raw_materials_data) + 2):
            rm_name = rm_sheet.cell(row=rm_row, column=1).value
            if rm_name == material_name:
                unit = rm_sheet.cell(row=rm_row, column=3).value
                bom_sheet.cell(row=bom_row_idx, column=3).value = unit
                # Unit Price
                price_cell = rm_sheet.cell(row=rm_row, column=2)
                bom_sheet.cell(row=bom_row_idx, column=5).value = f"='{rm_sheet.title}'!{price_cell.coordinate}"
                break
        # Total Quantity Required (Formula)
        qty_cell = bom_sheet.cell(row=bom_row_idx, column=2)
        qty_to_produce_cell = f"'Production Cost Calculator'!B{list(products.keys()).index(product_name)+2}"
        total_qty_formula = f"={qty_cell.coordinate}*{qty_to_produce_cell}"
        bom_sheet.cell(row=bom_row_idx, column=4).value = total_qty_formula
        # Total Cost (Formula)
        unit_price_cell = bom_sheet.cell(row=bom_row_idx, column=5)
        total_cost_formula = f"={bom_sheet.cell(row=bom_row_idx, column=4).coordinate}*{unit_price_cell.coordinate}"
        bom_sheet.cell(row=bom_row_idx, column=6).value = total_cost_formula

        bom_row_idx += 1

    # Calculate Total Cost for the Product
    total_cost_cell = bom_sheet.cell(row=bom_row_idx, column=6)
    total_cost_formula = f"=SUM(F2:F{bom_row_idx -1})"
    total_cost_cell.value = total_cost_formula
    bom_sheet.cell(row=bom_row_idx, column=5).value = "Total Cost:"

    # Link Total Cost back to Production Cost Calculator
    pcc_total_cost_cell = pcc_sheet.cell(row=list(products.keys()).index(product_name)+2, column=3)
    pcc_total_cost_cell.value = f"='{product_name}'!{total_cost_cell.coordinate}"
    # Cost per Unit (Formula)
    qty_to_produce_cell = pcc_sheet.cell(row=list(products.keys()).index(product_name)+2, column=2)
    cost_per_unit_formula = f"={pcc_total_cost_cell.coordinate}/{qty_to_produce_cell.coordinate}"
    pcc_sheet.cell(row=list(products.keys()).index(product_name)+2, column=4).value = cost_per_unit_formula

# Save the workbook
wb.save("production_costs.xlsx")
