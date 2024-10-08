import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.chart import LineChart, Reference
from openpyxl.writer.excel import save_virtual_workbook

# Define raw material costs
raw_material_costs = {
    'Дезин': 345,
    'Медихэнд': 46,
    'Дезискраб': 108,
    'Дезисепти ОП': 93,
    'Дезисепти Ультра': 72,
    'Дезихэнд': 35
}

# Define operational costs (these can be adjusted as needed)
operational_costs = {
    'Labor': 50000,
    'Utilities': 20000,
    'Rent': 30000,
    'Other': 10000
}

# Define production volumes
production_volumes = [15, 100, 500]  # in liters

# Create a new Excel workbook
wb = Workbook()

# Remove the default sheet
default_sheet = wb.active
wb.remove(default_sheet)

# Create 'Inputs' sheet
inputs_ws = wb.create_sheet('Inputs')

# Write raw material costs to 'Inputs' sheet
inputs_ws['A1'] = 'Product'
inputs_ws['B1'] = 'Raw Material Cost per Liter (₽)'
for idx, (product, cost) in enumerate(raw_material_costs.items(), start=2):
    inputs_ws[f'A{idx}'] = product
    inputs_ws[f'B{idx}'] = cost

# Write operational costs to 'Inputs' sheet
op_start_row = len(raw_material_costs) + 3
inputs_ws[f'A{op_start_row}'] = 'Operational Costs'
inputs_ws[f'A{op_start_row}'].font = Font(bold=True)
for idx, (op_cost, amount) in enumerate(operational_costs.items(), start=op_start_row + 1):
    inputs_ws[f'A{idx}'] = op_cost
    inputs_ws[f'B{idx}'] = amount

# Create 'Price Table' sheet
price_table_ws = wb.create_sheet('Price Table')

# Write headers
price_table_ws['A1'] = 'Product'
for idx, volume in enumerate(production_volumes, start=2):
    price_table_ws.cell(row=1, column=idx).value = f'Price for {volume}L (₽)'

# Calculate prices
for row_idx, product in enumerate(raw_material_costs.keys(), start=2):
    price_table_ws.cell(row=row_idx, column=1).value = product
    for col_idx, volume in enumerate(production_volumes, start=2):
        # Price = (Raw Material Cost * Volume) + Operational Costs per Unit
        raw_cost_cell = f"Inputs!$B${list(raw_material_costs.keys()).index(product)+2}"
        op_cost_total = sum(operational_costs.values())
        op_cost_per_unit = op_cost_total / (sum(production_volumes) * len(raw_material_costs))
        formula = f"=({raw_cost_cell}*{volume})+{op_cost_per_unit:.2f}"
        price_table_ws.cell(row=row_idx, column=col_idx).value = formula

# Format headers
for ws in [inputs_ws, price_table_ws]:
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

# Adjust column widths
for ws in wb.worksheets:
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

# Create 'Monthly Analytics' sheet
monthly_ws = wb.create_sheet('Monthly Analytics')
monthly_ws['A1'] = 'Month'
monthly_ws['B1'] = 'Total Revenue (₽)'
monthly_ws['C1'] = 'Total Expenses (₽)'
monthly_ws['D1'] = 'Net Profit (₽)'

# Sample data for 12 months
months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
          'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
for idx, month in enumerate(months, start=2):
    monthly_ws[f'A{idx}'] = month
    # Assuming some random values for demonstration
    monthly_ws[f'B{idx}'] = f"=SUM('Price Table'!B{2}:D{len(raw_material_costs)+1})*100"
    monthly_ws[f'C{idx}'] = sum(operational_costs.values())
    monthly_ws[f'D{idx}'] = f"=B{idx}-C{idx}"

# Create a line chart for Net Profit over months
chart = LineChart()
chart.title = "Net Profit Over Months"
data = Reference(monthly_ws, min_col=4, min_row=1, max_row=13)
cats = Reference(monthly_ws, min_col=1, min_row=2, max_row=13)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
monthly_ws.add_chart(chart, "F2")

# Create 'Overall Metrics' sheet
metrics_ws = wb.create_sheet('Overall Metrics')
metrics_ws['A1'] = 'Metric'
metrics_ws['B1'] = 'Value'

metrics_ws['A2'] = 'Total Revenue (Year)'
metrics_ws['B2'] = "=SUM('Monthly Analytics'!B2:B13)"

metrics_ws['A3'] = 'Total Expenses (Year)'
metrics_ws['B3'] = "=SUM('Monthly Analytics'!C2:C13)"

metrics_ws['A4'] = 'Net Profit (Year)'
metrics_ws['B4'] = "=B2-B3"

# Save the workbook
wb.save('business_analytics.xlsx')
